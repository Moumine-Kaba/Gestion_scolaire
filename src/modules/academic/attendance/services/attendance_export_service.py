# Service d'export et génération de rapports pour les présences
from typing import List, Dict, Optional
from datetime import datetime, timedelta
import pandas as pd
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from ..controllers.attendance_controller import AttendanceController
from ..controllers.attendance_stats_controller import AttendanceStatsController

class AttendanceExportService:
    """Service d'export et génération de rapports"""
    
    def __init__(self):
        self.attendance_controller = AttendanceController()
        self.stats_controller = AttendanceStatsController()
    
    def export_daily_attendance_pdf(self, classe_id: int, date: str, output_path: str) -> bool:
        """Exporte les présences du jour en PDF"""
        try:
            # Récupérer les données
            overview = self.attendance_controller.get_class_attendance_overview(classe_id, date)
            classe_name = self._get_classe_name(classe_id)
            
            # Créer le PDF
            pdf = FPDF()
            pdf.add_page()
            
            # En-tête
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(200, 10, txt=f"Feuille de Présence - {classe_name}", ln=True, align='C')
            pdf.cell(200, 8, txt=f"Date : {date}", ln=True, align='C')
            pdf.ln(10)
            
            # Statistiques
            stats = overview['stats']
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 8, txt="Statistiques du jour", ln=True)
            pdf.set_font("Arial", '', 10)
            pdf.cell(100, 6, f"Total élèves : {overview['total_students']}", ln=True)
            pdf.cell(100, 6, f"Présents : {stats.get('Présent', 0)}", ln=True)
            pdf.cell(100, 6, f"Absents : {stats.get('Absent', 0)}", ln=True)
            pdf.cell(100, 6, f"Retards : {stats.get('Retard', 0)}", ln=True)
            pdf.cell(100, 6, f"Justifiés : {stats.get('Justifié', 0)}", ln=True)
            pdf.ln(10)
            
            # Tableau des élèves
            pdf.set_font("Arial", 'B', 11)
            pdf.cell(40, 8, "Nom", 1)
            pdf.cell(40, 8, "Prénom", 1)
            pdf.cell(30, 8, "Statut", 1)
            pdf.cell(80, 8, "Commentaire", 1, ln=True)
            
            pdf.set_font("Arial", '', 9)
            for student in overview['students']:
                pdf.cell(40, 6, student['nom'], 1)
                pdf.cell(40, 6, student['prenom'], 1)
                pdf.cell(30, 6, student['statut'], 1)
                pdf.cell(80, 6, student['commentaire'][:50], 1, ln=True)
            
            # Pied de page
            pdf.ln(10)
            pdf.set_font("Arial", 'I', 8)
            pdf.cell(200, 6, txt=f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", ln=True, align='C')
            
            # Sauvegarde
            pdf.output(output_path)
            print(f"✅ PDF exporté : {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ Erreur export PDF: {e}")
            return False
    
    def export_monthly_attendance_excel(self, classe_id: int, month: int, year: int, output_path: str) -> bool:
        """Exporte les présences mensuelles en Excel"""
        try:
            # Récupérer les données mensuelles
            monthly_data = self.stats_controller.get_monthly_attendance_data(classe_id, year, month)
            classe_name = self._get_classe_name(classe_id)
            
            # Créer un DataFrame
            df = pd.DataFrame(monthly_data)
            
            # Créer le fichier Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Présences {month}/{year}"
            
            # En-tête
            ws['A1'] = f"Rapport de Présences - {classe_name}"
            ws['A2'] = f"Mois : {month}/{year}"
            ws['A3'] = f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            
            # Style de l'en-tête
            header_font = Font(bold=True, size=14)
            ws['A1'].font = header_font
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style du tableau
            self._style_excel_table(ws, len(df) + 5)
            
            # Sauvegarde
            wb.save(output_path)
            print(f"✅ Excel exporté : {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ Erreur export Excel: {e}")
            return False
    
    def export_student_attendance_history_pdf(self, eleve_id: int, output_path: str) -> bool:
        """Exporte l'historique complet d'un élève en PDF"""
        try:
            # Récupérer l'historique
            from ..controllers.attendance_history_controller import AttendanceHistoryController
            history_controller = AttendanceHistoryController()
            history = history_controller.get_student_history(eleve_id)
            
            if not history:
                print("❌ Aucun historique trouvé")
                return False
            
            student_name = f"{history[0]['prenom']} {history[0]['nom']}"
            
            # Créer le PDF
            pdf = FPDF()
            pdf.add_page()
            
            # En-tête
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(200, 10, txt=f"Historique des Présences", ln=True, align='C')
            pdf.cell(200, 8, txt=f"Élève : {student_name}", ln=True, align='C')
            pdf.ln(10)
            
            # Statistiques
            stats = self.stats_controller.get_student_attendance_stats(eleve_id)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 8, txt="Statistiques Globales", ln=True)
            pdf.set_font("Arial", '', 10)
            pdf.cell(100, 6, f"Total des jours : {stats.total_jours}", ln=True)
            pdf.cell(100, 6, f"Présences : {stats.presents}", ln=True)
            pdf.cell(100, 6, f"Absences : {stats.absents}", ln=True)
            pdf.cell(100, 6, f"Retards : {stats.retards}", ln=True)
            pdf.cell(100, 6, f"Justifiés : {stats.justifies}", ln=True)
            pdf.cell(100, 6, f"Taux de présence : {stats.taux_presence:.1f}%", ln=True)
            pdf.ln(10)
            
            # Tableau de l'historique
            pdf.set_font("Arial", 'B', 11)
            pdf.cell(40, 8, "Date", 1)
            pdf.cell(30, 8, "Statut", 1)
            pdf.cell(50, 8, "Classe", 1)
            pdf.cell(80, 8, "Commentaire", 1, ln=True)
            
            pdf.set_font("Arial", '', 9)
            for record in history:
                date_str = record['date'].strftime("%d/%m/%Y") if hasattr(record['date'], 'strftime') else str(record['date'])
                pdf.cell(40, 6, date_str, 1)
                pdf.cell(30, 6, record['statut'], 1)
                pdf.cell(50, 6, record['classe_nom'], 1)
                pdf.cell(80, 6, record['commentaire'][:50], 1, ln=True)
            
            # Sauvegarde
            pdf.output(output_path)
            print(f"✅ Historique PDF exporté : {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ Erreur export historique: {e}")
            return False
    
    def generate_attendance_summary_report(self, classe_id: int, start_date: str, end_date: str) -> Dict:
        """Génère un rapport de synthèse des présences"""
        try:
            # Récupérer le résumé
            summary = self.stats_controller.get_class_attendance_summary(classe_id, start_date, end_date)
            students = self.attendance_controller.get_students_by_class(classe_id)
            
            # Calculer les statistiques globales
            total_days = len(summary)
            total_possible = total_days * len(students) if total_days > 0 else 0
            
            total_presents = sum(day['presents'] for day in summary)
            total_absents = sum(day['absents'] for day in summary)
            total_retards = sum(day['retards'] for day in summary)
            total_justifies = sum(day['justifies'] for day in summary)
            
            overall_attendance_rate = (total_presents / total_possible * 100) if total_possible > 0 else 0
            
            # Identifier les élèves problématiques
            problematic_students = []
            for student in students:
                stats = self.stats_controller.get_student_attendance_stats(
                    student['id_eleve'], start_date, end_date
                )
                if stats.taux_presence < 80:  # Seuil de 80%
                    problematic_students.append({
                        'nom': f"{student['prenom']} {student['nom']}",
                        'taux_presence': stats.taux_presence,
                        'absents': stats.absents
                    })
            
            return {
                'periode': f"{start_date} à {end_date}",
                'total_days': total_days,
                'total_students': len(students),
                'total_presents': total_presents,
                'total_absents': total_absents,
                'total_retards': total_retards,
                'total_justifies': total_justifies,
                'overall_attendance_rate': overall_attendance_rate,
                'problematic_students': problematic_students,
                'daily_summary': summary
            }
            
        except Exception as e:
            print(f"❌ Erreur rapport synthèse: {e}")
            return {}
    
    def export_attendance_summary_pdf(self, summary_data: Dict, output_path: str) -> bool:
        """Exporte le rapport de synthèse en PDF"""
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # En-tête
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(200, 10, txt="Rapport de Synthèse des Présences", ln=True, align='C')
            pdf.cell(200, 8, txt=f"Période : {summary_data['periode']}", ln=True, align='C')
            pdf.ln(10)
            
            # Statistiques globales
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 8, txt="Statistiques Globales", ln=True)
            pdf.set_font("Arial", '', 10)
            pdf.cell(100, 6, f"Nombre de jours : {summary_data['total_days']}", ln=True)
            pdf.cell(100, 6, f"Nombre d'élèves : {summary_data['total_students']}", ln=True)
            pdf.cell(100, 6, f"Total présences : {summary_data['total_presents']}", ln=True)
            pdf.cell(100, 6, f"Total absences : {summary_data['total_absents']}", ln=True)
            pdf.cell(100, 6, f"Taux global : {summary_data['overall_attendance_rate']:.1f}%", ln=True)
            pdf.ln(10)
            
            # Élèves problématiques
            if summary_data['problematic_students']:
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(200, 8, txt="Élèves nécessitant une attention", ln=True)
                pdf.set_font("Arial", '', 10)
                
                for student in summary_data['problematic_students']:
                    pdf.cell(100, 6, f"{student['nom']} - {student['taux_presence']:.1f}% ({student['absents']} absences)", ln=True)
            
            # Sauvegarde
            pdf.output(output_path)
            print(f"✅ Rapport synthèse exporté : {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ Erreur export synthèse: {e}")
            return False
    
    def _get_classe_name(self, classe_id: int) -> str:
        """Récupère le nom d'une classe"""
        classes = self.attendance_controller.get_all_classes()
        for classe in classes:
            if classe['id_classe'] == classe_id:
                return classe['nom_classe']
        return "Classe inconnue"
    
    def _style_excel_table(self, ws, data_rows: int):
        """Applique un style au tableau Excel"""
        # Style de l'en-tête
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, 6):  # Colonnes A à E
            cell = ws.cell(row=5, column=col)  # Ligne 5 = en-tête des données
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
